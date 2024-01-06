import glob
import os.path

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

# 创建list名为val_eval_avg_acc_roc_auc_f1_gmean_best_record
val_eval_avg_acc_roc_auc_f1_gmean_best_record = []

# 用来存放处理后的DataFrame，用于最终写入Excel
sheets_dict = {}

# 假设CSV文件在当前目录的'csv_files'文件夹中
csv_files = glob.glob('./my_file/*.csv')

for i, file in enumerate(csv_files):
    # nrows = 50 if 'cr2' in file else 40  # TODO
    nrows = 100
    # nrows = 10000 if 'nll_pair_match' in file else nrows

    # （1）读取每个CSV文件的前50行（不包括header）
    df = pd.read_csv(file, nrows=nrows)

    # （2）在"val_eval_type2_acc"列后插入新列
    new_col = np.sqrt(df["val_eval_roc_auc"] * df["val_eval_gmean"])
    loc = df.columns.get_loc("val_eval_type2_acc") + 1
    df.insert(loc, "val_eval_sqrt_roc_auc_gmean", new_col)

    # （3）插入AVERAGE(val_eval_roc_auc, val_eval_gmean)
    df.insert(loc + 1, "val_eval_avg_roc_auc_gmean", (df["val_eval_roc_auc"] + df["val_eval_gmean"]) / 2)

    # （4）插入AVERAGE(val_eval_acc, val_eval_roc_auc, val_eval_gmean)
    df.insert(loc + 2, "val_eval_avg_acc_roc_auc_gmean",
              (df["val_eval_acc"] + df["val_eval_roc_auc"] + df["val_eval_gmean"]) / 3)

    # （5）插入AVERAGE(val_eval_acc, val_eval_roc_auc, val_eval_f1, val_eval_gmean)
    df.insert(loc + 3, "val_eval_avg_acc_roc_auc_f1_gmean",
              (df["val_eval_acc"] + df["val_eval_roc_auc"] +
               df["val_eval_f1"] + df["val_eval_gmean"]) / 4)

    # （6）找出val_eval_avg_acc_roc_auc_f1_gmean_best_record列中最大的值所在行
    best_record = df.loc[df["val_eval_avg_acc_roc_auc_f1_gmean"].idxmax()]
    val_eval_avg_acc_roc_auc_f1_gmean_best_record.append(best_record)

    # 将处理后的DataFrame存储起来，稍后写入Excel
    # sheet_name = os.path.basename(file)
    sheet_name = str(i)
    sheets_dict[sheet_name] = df

# hidden_columns = [
#     "val_eval_threshold", "val_eval_pr_auc", "val_eval_recall", "val_eval_precision",
#     "val_eval_tn", "val_eval_fp", "val_eval_fn", "val_eval_tp", "val_eval_cm",
#     "val_eval_cpi", "val_eval_runtime", "val_eval_samples_per_second", "val_eval_steps_per_second",
#     "val_epoch", "test_eval_loss", "test_eval_threshold", "test_eval_pr_auc", "test_eval_recall",
#     "test_eval_precision", "test_eval_tn", "test_eval_fp", "test_eval_fn", "test_eval_tp",
#     "test_eval_cm", "test_eval_cpi", "test_eval_runtime", "test_eval_samples_per_second",
#     "test_eval_steps_per_second", "test_epoch"
# ]
# 创建一个新的Excel文件名为results
with pd.ExcelWriter('processed.xlsx', engine='openpyxl') as writer:
    # 将每个处理后的DataFrame写入一个sheet
    for sheet_name, df in sheets_dict.items():
        # 写入DataFrame到Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        # 获取当前sheet
        worksheet = writer.sheets[sheet_name]
        # # 调整列宽
        # for column in df.columns:
        #     column_length = max(df[column].astype(str).map(len).max(), len(column))
        #     col_idx = df.columns.get_loc(column) + 1  # pandas是基于0的索引，openpyxl是基于1
        #     worksheet.column_dimensions[get_column_letter(col_idx)].width = column_length
        # # 隐藏指定的列
        # for column in hidden_columns:
        #     if column in df.columns:
        #         col_idx = df.columns.get_loc(column) + 1
        #         worksheet.column_dimensions[get_column_letter(col_idx)].hidden = True

    # 将eval_final_metric_best_record写入到一个新的sheet中
    df_eval_final = pd.DataFrame(val_eval_avg_acc_roc_auc_f1_gmean_best_record)
    df_eval_final.to_excel(writer, sheet_name='Best Records', index=False)
    # # 调整'Best Records' sheet的列宽并隐藏指定列
    # best_records_worksheet = writer.sheets['Best Records']
    # for column in df_eval_final.columns:
    #     column_length = max(df_eval_final[column].astype(str).map(len).max(), len(column))
    #     col_idx = df_eval_final.columns.get_loc(column) + 1
    #     best_records_worksheet.column_dimensions[get_column_letter(col_idx)].width = column_length
    #     # 隐藏指定的列
    #     if column in hidden_columns:
    #         best_records_worksheet.column_dimensions[get_column_letter(col_idx)].hidden = True

# 如果需要确认输出，可以打印eval_final_metric_best_record
# print(eval_final_metric_best_record)
