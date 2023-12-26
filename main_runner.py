import glob
import json
import os
import random
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd
import torch
import torchinfo
from torch.utils.data import ConcatDataset
from transformers import AutoTokenizer, Trainer, EarlyStoppingCallback, HfArgumentParser, TrainerCallback, \
    PreTrainedTokenizerFast
from transformers.utils import logging

from crmm import runner_setup
from crmm.arguments import LanguageModelArguments, MultimodalDataArguments, CrmmTrainingArguments
from crmm.dataset.multimodal_data import MultimodalData
from crmm.dataset.multimodal_dataset import MultimodalCollator
from crmm.metrics import calc_classification_metrics_hf
from crmm.models import runner_dict
from crmm.models.clncp import CLNCPConfig, CLNCP, TASK_MODE_DICT
from crmm.runner_callback import ContrastiveEarlyStoppingCallBack, EvaluateEveryNthEpoch

logger = logging.get_logger('transformers')

os.environ['COMET_MODE'] = 'DISABLED'
os.environ["TOKENIZERS_PARALLELISM"] = "true"
os.environ["TRANSFORMERS_OFFLINE"] = "1"
# os.environ["WANDB_DISABLED"] = "true"

# 设置随机种子
# seed = 3407
seed = random.randint(0, 10000)
torch.manual_seed(seed)
np.random.seed(seed)
random.seed(seed)


def main(language_model_args: LanguageModelArguments,
         data_args: MultimodalDataArguments,
         training_args: CrmmTrainingArguments):
    # @@@@ 1. SETUP
    training_args.seed = seed
    task = training_args.task

    # @@@@ 2. DATASET
    mm_data = MultimodalData(data_args)
    train_dataset, test_dataset, val_dataset = mm_data.get_datasets()
    n_labels = len(mm_data.label_values)
    if task == 'pretrain':
        # if pretrain task, concat train and val to unsupervised train!
        train_dataset = ConcatDataset([train_dataset, val_dataset])
        val_dataset = None

    # @@@@ 3. TOKENIZER AND MODEL
    text_tokenizer = AutoTokenizer.from_pretrained(
        pretrained_model_name_or_path=language_model_args.language_model_name,
        cache_dir=language_model_args.cache_dir,
        local_files_only=language_model_args.load_hf_model_from_cache
    ) if 'text' in training_args.use_modality or 'text' in training_args.fuse_modality else None
    cat_tokenizer = PreTrainedTokenizerFast(
        tokenizer_file=f"data/{data_args.dataset_name}/{data_args.dataset_name}-bpe.tokenizer.json",
        pad_token="[PAD]",
        pad_token_id=3,
        cls_token="[CLS]",
        cls_token_id=1,
        sep_token="[SEP]",
        sep_token_id=2,
        unk_token="[UNK]",
        unk_token_id=0,
        mask_token="[MASK]",
        mask_token_id=4
    ) if 'cat' in training_args.use_modality or 'cat' in training_args.fuse_modality else None
    default_model_config = CLNCPConfig(
        n_labels=n_labels,
        num_feat_dim=test_dataset.numerical_feats.shape[1],
        use_modality=training_args.use_modality,
        fuse_modality=training_args.fuse_modality,
        modality_fusion_method=training_args.modality_fusion_method,
        max_text_length=language_model_args.max_seq_length,
        language_model_cache_dir=language_model_args.cache_dir,
        language_model_name=language_model_args.language_model_name,
        language_model_local_files_only=language_model_args.load_hf_model_from_cache,
        load_hf_pretrained=language_model_args.load_hf_pretrained,
        freeze_language_model_params=language_model_args.freeze_language_model_params,
        mode=TASK_MODE_DICT[task],
        contrastive_targets=training_args.contrastive_targets,
        cat_vocab_size=cat_tokenizer.vocab_size if cat_tokenizer is not None else None
    )
    if task == 'pretrain':
        if training_args.pretrained_model_dir is None:
            model = CLNCP(default_model_config)
        else:
            # i.e., continue pretraining
            model = CLNCP.from_pretrained(training_args.pretrained_model_dir)
    elif task == 'pair_match_evaluation':
        # model_config = CLNCPConfig.from_pretrained(training_args.pretrained_model_dir)
        # model_config.mode = TASK_MODE_DICT[task]  # should update to match the task
        # model = CLNCP.from_pretrained(training_args.pretrained_model_dir, config=model_config)
        model = CLNCP.from_pretrained(training_args.pretrained_model_dir, config=default_model_config)
    elif task == 'finetune_classification':
        # model_config = CLNCPConfig.from_pretrained(training_args.pretrained_model_dir)
        # # note: freeze_language_model_params can be overridden by training_args.freeze_language_model_params in run command!
        # model_config.freeze_language_model_params = language_model_args.freeze_language_model_params
        # model_config.mode = TASK_MODE_DICT[task]  # should update to match the task
        # model = CLNCP.from_pretrained(training_args.pretrained_model_dir, config=model_config)
        model = CLNCP.from_pretrained(training_args.pretrained_model_dir, config=default_model_config)
        # model.feature_extractors['num'].requires_grad_(False)
        # model.feature_extractors['cat'].requires_grad_(False)
        # model.feature_extractors['joint'].requires_grad_(False)
    elif task == 'finetune_classification_scratch':
        model = CLNCP(default_model_config)
    elif task == 'finetune_classification_evaluation':
        model = CLNCP.from_pretrained(training_args.finetuned_model_dir)
    elif task == 'multi_task_classification':
        if training_args.pretrained_model_dir is None:
            model = CLNCP(default_model_config)
        else:
            model_config = CLNCPConfig.from_pretrained(training_args.pretrained_model_dir)
            # note: freeze_language_model_params can be overridden by training_args.freeze_language_model_params in run command!
            model_config.freeze_language_model_params = language_model_args.freeze_language_model_params
            model = CLNCP.from_pretrained(training_args.pretrained_model_dir, config=model_config)
    # elif task == 'zero_shot_prediction_from_another_dataset_pretraining':
    #     # load model pretrained on another dataset
    #     pretrained_model_config = CLNCPConfig.from_pretrained(training_args.pretrained_model_dir)
    #     pretrained_model = CLNCP.from_pretrained(training_args.pretrained_model_dir, config=pretrained_model_config)
    #     # create model for current dataset
    #     model_config = CLNCPConfig(**default_model_config_params,
    #                                freeze_language_model_params=language_model_args.freeze_language_model_params,
    #                                pretrained=False)
    #     model = CLNCP(model_config)
    #     model.pretrained = True  # should manually set to pretrained
    #     # use pretrained clip text weights
    #     model.feature_extractors['text'] = pretrained_model.feature_extractors['text']
    # elif task in ['conventional_multimodal_classification', 'no_clip_classification']:
    #     model_config = CMMConfig(**default_model_config_params)
    #     model = ConventionalMultimodalClassification(model_config)
    else:
        raise ValueError(f'Unknown task: {task}')
    model_config = model.config
    logger.info(f'\n{model}')
    logger.info(f'\n{torchinfo.summary(model, verbose=0)}')

    # @@@@ 4. TRAINING
    callbacks = [
        # EarlyStoppingCallback(training_args.patience),
        ContrastiveEarlyStoppingCallBack(training_args.contrastive_early_stopping_epoch),
        EvaluateEveryNthEpoch(val_dataset, test_dataset, n_epochs=1, csv_path=training_args.save_hist_eval_csv_path)
    ] if 'classification' in task else None
    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=train_dataset,
        eval_dataset=test_dataset if 'evaluation' in task else val_dataset,  # can be None in pretrain
        compute_metrics=calc_classification_metrics_hf if 'evaluation' in task or 'classification' in task else None,
        callbacks=callbacks,
        data_collator=MultimodalCollator(
            text_tokenizer,
            cat_tokenizer,
            language_model_args.max_seq_length,
            data_args.natural_language_labels,
        ),
    )
    if callbacks is not None:
        callbacks[1].trainer = trainer

    if 'evaluation' not in task:
        trainer.train()
        trainer.save_model()  # manually save at end

    best_model_checkpoint = trainer.state.best_model_checkpoint
    best_step = int(best_model_checkpoint.split("-")[-1]) if best_model_checkpoint else None
    total_step = trainer.state.max_steps
    if 'evaluation' not in task:
        logger.info(f"Best model path: {best_model_checkpoint}")
        logger.info(f"Best metric value: {trainer.state.best_metric}")

    # @@@@ 5. EVALUATION
    if 'evaluation' in task or 'classification' in task:
        val_best_results = trainer.evaluate(eval_dataset=val_dataset)
        test_results = trainer.evaluate(eval_dataset=test_dataset)
        tasks_need_pretrain = ['pair_match_evaluation', 'finetune_classification']
        # if task in tasks_need_pretrain + ['zero_shot_prediction_from_another_dataset_pretraining']:
        if task in tasks_need_pretrain:
            with open(os.path.join(training_args.pretrained_model_dir, 'training_arguments.json')) as file:
                training_arguments = json.load(file)

            ckpt_dir = glob.glob(f'{training_args.pretrained_model_dir}/*/')[0]
            with open(os.path.join(ckpt_dir, 'trainer_state.json')) as file:
                trainer_state = json.load(file)

            pretrain_final_step = trainer_state['global_step']
            pretrain_max_steps = trainer_state['max_steps']
            pretrain_epoch = training_arguments['num_train_epochs']
            pretrain_model_final_ckpt_epoch = pretrain_final_step / pretrain_max_steps * pretrain_epoch
            pretrain_steps_per_epoch = pretrain_max_steps / pretrain_epoch

        basic_info = {
            'dataset': data_args.dataset_name,
            'dataset_info': data_args.dataset_info,
            'data_path': data_args.data_path,
            'task': task,
            'language_model': language_model_args.language_model_name,
            'freeze_language_model_params': model_config.freeze_language_model_params,
            'numerical': 'num' in model_config.use_modality,
            'category': 'cat' in model_config.use_modality,
            'text': 'text' in model_config.use_modality,
            'modality_fusion_method': model_config.modality_fusion_method,
            'this_exp_path': training_args.output_dir,
            'natural_language_labels': data_args.natural_language_labels if task == 'pair_match_evaluation' else None,
            'pretrain_model_path': training_args.pretrained_model_dir if task in tasks_need_pretrain else None,
            'pretrain_epoch': pretrain_epoch if task in tasks_need_pretrain else None,
            'pretrain_model_final_ckpt_epoch': pretrain_model_final_ckpt_epoch if task in tasks_need_pretrain else None,
            'pretrain_steps_per_epoch': pretrain_steps_per_epoch if task in tasks_need_pretrain else None,
            'classification_train_best_step': None if task == 'pair_match_evaluation' else best_step,
            'classification_train_total_step': None if task == 'pair_match_evaluation' else total_step,
            'classification_train_batch_size': None if task == 'pair_match_evaluation' else training_args.per_device_train_batch_size,
            'classification_train_epoch': None if task == 'pair_match_evaluation' else training_args.num_train_epochs,
        }

        save_excel(val_best_results, test_results, basic_info, training_args.save_excel_path)


def save_excel(val_best_results, test_results, basic_info, excel_path):
    logger.info(f'** save results to {excel_path}')
    val_data = {f'val_{k}': v for k, v in val_best_results.items()}
    test_data = {f'test_{k}': v for k, v in test_results.items()}

    if os.path.exists(excel_path):
        book = openpyxl.load_workbook(excel_path)
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        writer._book = book

        if 'Sheet1' in book.sheetnames:
            startrow = writer.sheets['Sheet1'].max_row
        else:
            startrow = 0

    else:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl')
        startrow = 0

    data = {**{'Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}, **basic_info, **val_data, **test_data, }
    # 第一部分所需的键
    keys_part1 = [
        'Timestamp', 'dataset', 'task', 'val_eval_acc', 'val_eval_roc_auc',
        'val_eval_ks', 'val_eval_gmean', 'val_eval_type1_acc', 'val_eval_type2_acc',
        'test_eval_acc', 'test_eval_roc_auc', 'test_eval_ks', 'test_eval_gmean',
        'test_eval_type1_acc', 'test_eval_type2_acc'
    ]

    # 创建两个新的字典，分别存储两部分
    part1 = {key: data[key] for key in keys_part1}
    part2 = {key: data[key] for key in data if key not in keys_part1}

    # 将第二部分转换成JSON字符串
    json_part2 = json.dumps(part2, indent=4)

    df = pd.DataFrame([{**part1, 'other_info': json_part2}])

    df.to_excel(writer, index=False, header=(startrow == 0), startrow=startrow)

    writer._save()
    writer.close()


if __name__ == '__main__':
    parser = HfArgumentParser([LanguageModelArguments, MultimodalDataArguments, CrmmTrainingArguments])
    _bert_model_args, _data_args, _training_args = parser.parse_args_into_dataclasses()
    _training_args = runner_setup.setup(_training_args, _data_args, _bert_model_args)

    main(_bert_model_args, _data_args, _training_args)
